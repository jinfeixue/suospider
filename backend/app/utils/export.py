"""Data export utilities - Excel and JSON."""
import json
import os
from datetime import datetime
from typing import List, Dict, Any

from app.config import settings


def export_json(data: List[Dict[str, Any]], filename: str = None) -> str:
    """Export data to JSON file. Returns file path."""
    if filename is None:
        filename = f"export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
    filepath = os.path.join(settings.RESULTS_DIR, filename)
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    return filepath


def export_excel(data: List[Dict[str, Any]], filename: str = None) -> str:
    """Export data to Excel file. Returns file path."""
    from openpyxl import Workbook

    if filename is None:
        filename = f"export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(settings.RESULTS_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    if not data:
        wb.save(filepath)
        return filepath

    # Write headers
    headers = list(data[0].keys())
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=str(header))

    # Write data rows
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            val = row_data.get(header, "")
            if isinstance(val, (dict, list)):
                val = json.dumps(val, ensure_ascii=False)
            ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else "")

    wb.save(filepath)
    return filepath
